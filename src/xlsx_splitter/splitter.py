import openpyxl
from copy import copy


class ExcelSplitter:

    def __init__(self, filename):
        self.ws = openpyxl.load_workbook(filename=filename)

    def split_by_row_in_average(self, sheet_name, num, target_file, save_in_one_file=True):
        if not save_in_one_file:
            if target_file.endswith(".xlsx"):
                import os
                target_file = target_file.rsplit(".xlsx", 1)[0]
                if not os.path.exists(target_file):
                    os.mkdir(target_file)
        # load source sheet
        res_sheets = []
        _sheet = self.ws[sheet_name]
        rows = [r for r in _sheet.iter_rows()]
        
        # split the sheet in average
        _a, _b = divmod(len(rows), num)
        row_splits = [rows[i*_a+min(i, _b) : (i+1)*_a+min(i+1, _b)] for i in range(num)]
        if save_in_one_file:
            _wb = openpyxl.Workbook()
        else:
            _wbs = {}

        for idx, row_split in enumerate(row_splits):
            start_row_id = row_split[0][0].row
            if save_in_one_file:
                _ws = _wb.create_sheet(f"{sheet_name}_split_{idx}")
            else:
                _wb = openpyxl.Workbook()
                _wbs[os.path.join(target_file, f"split_{idx}.xlsx")] = _wb
                _ws = _wb.active
            for row in row_split:
                for cell in row:
                    _cell = _ws.cell(row=cell.row-start_row_id+1, column=cell.column, value=cell.value)
                    if cell.has_style:
                        _cell.font = copy(cell.font)
                        _cell.border = copy(cell.border)
                        _cell.fill = copy(cell.fill)
                        _cell.number_format = copy(cell.number_format)
                        _cell.protection = copy(cell.protection)
                        _cell.alignment = copy(cell.alignment)

        if save_in_one_file:
            _wb.remove(_wb[_wb.sheetnames[0]])
            _wb.save(target_file)
        else:
            for k,v in _wbs.items():
                v.save(k)


if __name__ == "__main__":
    input_file = "./test.xlsx"
    output_file = input_file.rsplit('.', 1)[0]+"_split.xlsx"
    ExcelSplitter(input_file).split_by_row_in_average("Sheet1", 5, output_file, save_in_one_file=False)
